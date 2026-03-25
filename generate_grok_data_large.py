#!/usr/bin/env python3
"""
Adeptus Mechanicus Prayer Dataset Generator — Excel Edition

Changes from original:
  - Components and operations loaded from mechanicus_components.xlsx
  - Tracks (component, operation) pairs already generated in JSON batch files
  - Skips pairs that already meet TARGET_PER_PAIR — no re-spending money
  - Writes pair coverage summary back to the Excel Pair Log sheet after run
"""
import json
import time
import hashlib
import random
import re
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Set
from collections import defaultdict
import os

try:
    import pandas as pd
except ImportError:
    print("ERROR: pip install pandas openpyxl")
    exit(1)

try:
    from openai import OpenAI
except ImportError:
    print("ERROR: pip install openai")
    exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl")
    exit(1)

from dotenv import load_dotenv
load_dotenv()

# ============================================================================
# CONFIGURATION
# ============================================================================
XAI_API_KEY = os.getenv("XAI_API_KEY")
if not XAI_API_KEY:
    print("ERROR: Set XAI_API_KEY environment variable!")
    exit(1)

MODEL        = "grok-4-1-fast-reasoning"
XAI_BASE_URL = "https://api.x.ai/v1"

# ── Paths ────────────────────────────────────────────────────────────────────
EXCEL_FILE  = Path("mechanicus_components.xlsx")   # component/operation source
OUTPUT_DIR  = Path("mechanicus_dataset_grok")       # batch JSON files land here

# ── Generation knobs ─────────────────────────────────────────────────────────
TARGET_PER_PAIR    = 2      # how many examples we want per (component, operation) pair
BATCH_SIZE         = 10     # items sent per API call
RATE_LIMIT_DELAY   = 1.0
TEMPERATURE        = 0.8
MAX_TOKENS         = 3000
MAX_RETRIES        = 3

# ── Validation ───────────────────────────────────────────────────────────────
MIN_PRAYER_LINES = 3
MAX_PRAYER_LINES = 7

# ============================================================================
# LOAD COMPONENTS & OPERATIONS FROM EXCEL
# ============================================================================
def load_excel_data(path: Path) -> Tuple[List[str], List[str]]:
    """Read enabled components and operations from the workbook."""
    if not path.exists():
        print(f"ERROR: Excel file not found: {path}")
        print("  Run create_excel.py first to generate mechanicus_components.xlsx")
        exit(1)

    df_comp = pd.read_excel(path, sheet_name="Components", header=2)
    df_ops  = pd.read_excel(path, sheet_name="Operations",  header=1)

    # Components — keep rows where Enabled == True and Component column is non-null
    comp_col    = df_comp.columns[1]   # "Component"
    enabled_col = df_comp.columns[4]   # "Enabled"

    def is_enabled(series):
        """Accept True, 1, 1.0, 'true', 'yes' — handles Excel's float booleans."""
        def _check(v):
            if v is None:
                return False
            if isinstance(v, (bool, int, float)):
                return bool(v)
            return str(v).strip().lower() in ("true", "yes", "1")
        return series.apply(_check)

    enabled_mask = is_enabled(df_comp[enabled_col])
    components = (
        df_comp.loc[enabled_mask, comp_col]
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )
    # Drop separator / header-like rows that snuck in
    components = [c for c in components if c and not c.startswith("▼") and c != comp_col]

    # Operations
    op_col      = df_ops.columns[1]   # "Operation"
    op_enabled  = df_ops.columns[2]   # "Enabled"

    op_enabled_mask = is_enabled(df_ops[op_enabled])
    operations = (
        df_ops.loc[op_enabled_mask, op_col]
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )
    operations = [o for o in operations if o and o != op_col]

    print(f"📋 Loaded from Excel:")
    print(f"   Components: {len(components)}")
    print(f"   Operations: {len(operations)}")
    return components, operations


# ============================================================================
# COUNT EXISTING PAIRS FROM BATCH FILES
# ============================================================================
def count_existing_pairs(output_dir: Path) -> Dict[Tuple[str, str], int]:
    """
    Scan all batch_*.json files and count how many examples exist
    for every (component, operation) pair.
    """
    pair_counts: Dict[Tuple[str, str], int] = defaultdict(int)

    for batch_file in sorted(output_dir.glob("batch_*.json")):
        try:
            with open(batch_file, encoding="utf-8") as f:
                items = json.load(f)
            for item in items:
                comp = item.get("component", "").strip()
                op   = item.get("operation", "").strip()
                if comp and op:
                    pair_counts[(comp, op)] += 1
        except Exception as e:
            print(f"  ⚠ Could not read {batch_file.name}: {e}")

    return pair_counts


def build_todo_list(
    components: List[str],
    operations: List[str],
    pair_counts: Dict[Tuple[str, str], int],
    target: int,
) -> List[Tuple[str, str, int]]:
    """
    Return a list of (component, operation, needed) for pairs that still need
    more examples.  Sorted so rarest pairs are tackled first.
    """
    todo = []
    for comp in components:
        for op in operations:
            have   = pair_counts.get((comp, op), 0)
            needed = target - have
            if needed > 0:
                todo.append((comp, op, needed))

    # Rarest first → fastest way to hit full coverage
    todo.sort(key=lambda x: x[2], reverse=True)
    return todo


# ============================================================================
# PROMPT TEMPLATES
# ============================================================================
PROMPT_TEMPLATES = [
    "Help me {operation} the {component}",
    "I need to {operation} the {component}",
    "Guide me through {component} {operation}",
    "Assist with {component} {operation}",
    "The {component} needs {operation}",
    "The {component} requires {operation}",
    "My {component} is ready for {operation}",
    "Emergency: {component} {operation}",
    "Urgent {component} {operation} needed",
    "Critical: {component} requires {operation}",
    "Bless the {component} for {operation}",
    "Sanctify the {component} during {operation}",
    "Consecrate the {component} before {operation}",
    "Offer prayer for {component} {operation}",
    "What's the prayer for {component} {operation}?",
    "How do I bless the {component} for {operation}?",
    "Can you guide {component} {operation}?",
    "Prepare the {component} for {operation}",
    "Ready the {component} for {operation}",
    "The {component} awaits {operation}",
]

SYSTEM_PROMPT = """You are a prayer generator for Warhammer 40,000 Adeptus Mechanicus tech-priests.
CRITICAL REQUIREMENTS:
1. OUTPUT FORMAT: Return ONLY valid JSON. No markdown, no explanation, no preamble.
   Must start with { and end with }
2. PRAYER STRUCTURE:
   - Write 3-6 lines of natural flowing prose (NO "First:", "Second:", "Third:" labels)
   - Each line describes a sequential ritual step
   - Use archaic/Latinate language mixed with technical terminology
   - Include the COMPONENT name at least twice
   - Reference the OPERATION at least once
   - Mention Machine Spirits and/or the Omnissiah appropriately
3. STYLE EXAMPLES:
BAD (has labels):
First: Approach the cogitator with reverence.
Second: Clear corrupted data fragments.
GOOD (natural prose):
Approach the cogitator with reverence and speak the Litany of Awakening.
Clear all corrupted data-fragments from its blessed memory banks.
Initiate the boot sequence whilst anointing each circuit with sacred oil.
Verify the Machine Spirit's return through diagnostic incantations.
4. RETURN FORMAT:
{"prayers": [{"id": "string", "prayer": "string with real newlines between lines"}]}
5. LANGUAGE: Archaic verbs (anoint, invoke, beseech, sanctify), tech-religious fusion,
   Machine Spirit references, Omnissiah invocations. NO modern casual language."""


# ============================================================================
# DEDUPLICATION
# ============================================================================
class DeduplicationTracker:
    def __init__(self):
        self.seen: Set[str] = set()

    def _key(self, prompt: str) -> str:
        return hashlib.md5(prompt.lower().strip().encode()).hexdigest()

    def is_duplicate(self, prompt: str) -> bool:
        return self._key(prompt) in self.seen

    def add(self, prompt: str):
        self.seen.add(self._key(prompt))


# ============================================================================
# VALIDATION
# ============================================================================
def normalize_prayer(prayer: str) -> str:
    prayer = prayer.replace("\\n", "\n")
    lines = [l.strip() for l in prayer.split("\n") if l.strip()]
    return "\n".join(lines)


def validate_prayer(prayer: str, component: str, operation: str) -> Tuple[bool, List[str]]:
    reasons = []
    prayer  = normalize_prayer(prayer)
    lines   = [l for l in prayer.split("\n") if l.strip()]

    if not (MIN_PRAYER_LINES <= len(lines) <= MAX_PRAYER_LINES):
        reasons.append(f"line_count={len(lines)}")

    if re.search(r'\b(First|Second|Third|Fourth|Fifth):', prayer):
        reasons.append("has_step_labels")

    p_lower = prayer.lower()
    if component.lower() not in p_lower:
        reasons.append("component_missing")

    op_words = {operation.lower(), operation.lower().rstrip("s"),
                operation.lower() + "ed", operation.lower() + "ing"}
    if not any(w in p_lower for w in op_words):
        reasons.append("operation_missing")

    flavor = ["omnissiah", "machine spirit", "machine-spirit", "blessed", "sacred",
              "holy", "litany", "canticle", "rite", "ritual", "consecrate",
              "sanctify", "invoke", "anoint", "prayer"]
    if not any(f in p_lower for f in flavor):
        reasons.append("no_mechanicus_flavor")

    return (len(reasons) == 0), reasons


# ============================================================================
# API
# ============================================================================
def call_grok(client: OpenAI, specs: List[Dict]) -> List[Dict]:
    items = [{"id": s["id"], "component": s["component"],
              "operation": s["operation"], "user_prompt": s["user_prompt"]}
             for s in specs]

    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content":
             "Generate prayers for these requests. Return ONLY the JSON.\n\n"
             "ITEMS:\n" + json.dumps(items, indent=2)},
        ],
        temperature=TEMPERATURE,
        max_tokens=MAX_TOKENS,
    )

    text = resp.choices[0].message.content.strip()
    if "```" in text:
        parts = text.split("```")
        text  = max(parts, key=lambda s: s.count("{") + s.count("}")).strip()
        if text.startswith("json"):
            text = text[4:].strip()

    start = text.find("{"); end = text.rfind("}") + 1
    if start == -1 or end <= 0:
        raise ValueError("No JSON in response")

    data = json.loads(text[start:end])
    if "prayers" not in data:
        raise ValueError("Missing 'prayers' key")
    return data["prayers"]


# ============================================================================
# BATCH GENERATION
# ============================================================================
def make_prompt(component: str, operation: str) -> str:
    template = random.choice(PROMPT_TEMPLATES)
    if any(kw in template for kw in ["for {operation}", "during {operation}", "before {operation}"]):
        return template.format(component=component, operation=operation)
    vowel = operation[0].lower() in "aeiou"
    op_with_art = ("an " if vowel else "a ") + operation
    return template.format(component=component, operation=op_with_art)


def generate_batch(client, specs_batch: List[Dict], dedup: DeduplicationTracker) -> List[Dict]:
    spec_by_id = {s["id"]: s for s in specs_batch}

    for attempt in range(MAX_RETRIES):
        try:
            raw = call_grok(client, specs_batch)
            by_id = {r["id"]: r for r in raw if isinstance(r, dict) and "id" in r}

            results, failed = [], []
            for sid, spec in spec_by_id.items():
                item = by_id.get(sid)
                if not item:
                    failed.append((sid, ["missing_from_response"]))
                    continue

                prayer = normalize_prayer(item.get("prayer", ""))
                ok, reasons = validate_prayer(prayer, spec["component"], spec["operation"])

                if ok:
                    results.append({
                        "user_prompt": spec["user_prompt"],
                        "prayer":      prayer,
                        "component":   spec["component"],
                        "operation":   spec["operation"],
                        "spec_id":     sid,
                    })
                else:
                    failed.append((sid, reasons))

            if failed:
                print(f"    ⚠ {len(failed)} failed: "
                      + ", ".join(f"{sid}:{r}" for sid, r in failed[:3]))

            if results:
                return results

            if attempt < MAX_RETRIES - 1:
                time.sleep(RATE_LIMIT_DELAY)

        except Exception as e:
            print(f"    ✗ API error (attempt {attempt+1}): {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(RATE_LIMIT_DELAY * 2)

    return []


# ============================================================================
# EXCEL PAIR LOG UPDATE
# ============================================================================
def write_pair_log(excel_path: Path, pair_counts: Dict, components: List[str],
                   operations: List[str], target: int):
    """Overwrite the Pair Log sheet with current coverage."""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Pair Log"]

        # Clear rows below header (row 2) and note (row 3)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        DONE   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
        TODO   = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
        ALT    = PatternFill("solid", start_color="EFEFEF", end_color="EFEFEF")

        row = 4
        for i, comp in enumerate(components):
            for op in operations:
                count  = pair_counts.get((comp, op), 0)
                status = "✓ done" if count >= target else f"need {target - count}"
                bg     = DONE if count >= target else (TODO if i % 2 == 0 else ALT)

                for col, val in enumerate([comp, op, count, target, status], 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font      = Font(name="Arial", size=9)
                    c.fill      = bg
                    c.border    = border
                    c.alignment = Alignment(vertical="center")
                row += 1

        wb.save(excel_path)
        print(f"✓ Pair Log updated in {excel_path}")
    except Exception as e:
        print(f"⚠ Could not update Pair Log: {e}")


# ============================================================================
# MAIN
# ============================================================================
def main():
    client = OpenAI(api_key=XAI_API_KEY, base_url=XAI_BASE_URL)
    OUTPUT_DIR.mkdir(exist_ok=True)

    # ── Load config from Excel ────────────────────────────────────────────────
    components, operations = load_excel_data(EXCEL_FILE)

    # ── Count what we already have ────────────────────────────────────────────
    pair_counts = count_existing_pairs(OUTPUT_DIR)
    total_existing = sum(pair_counts.values())

    # ── Build todo list ───────────────────────────────────────────────────────
    todo = build_todo_list(components, operations, pair_counts, TARGET_PER_PAIR)
    total_needed = sum(n for _, _, n in todo)
    pairs_done   = sum(1 for c in components for o in operations
                       if pair_counts.get((c, o), 0) >= TARGET_PER_PAIR)
    pairs_total  = len(components) * len(operations)

    print("\n🔧 Adeptus Mechanicus Prayer Dataset Generator (Excel + Pair-Aware)")
    print("=" * 70)
    print(f"Components:         {len(components)}")
    print(f"Operations:         {len(operations)}")
    print(f"Target per pair:    {TARGET_PER_PAIR}")
    print(f"Total pairs:        {pairs_total:,}")
    print(f"Pairs complete:     {pairs_done:,}  ({100*pairs_done/pairs_total:.1f}%)")
    print(f"Pairs remaining:    {len(todo):,}")
    print(f"Examples existing:  {total_existing:,}")
    print(f"Examples needed:    {total_needed:,}")
    print(f"Approx API calls:   ~{(total_needed + BATCH_SIZE - 1) // BATCH_SIZE:,}")
    print("=" * 70)

    if not todo:
        print("\n✅ All pairs already meet the target — nothing to generate!")
        write_pair_log(EXCEL_FILE, pair_counts, components, operations, TARGET_PER_PAIR)
        return

    # ── Generation loop ───────────────────────────────────────────────────────
    dedup      = DeduplicationTracker()
    all_new    = []
    batch_num  = max((int(re.search(r'\d+', f.stem).group())
                      for f in OUTPUT_DIR.glob("batch_*.json")
                      if re.search(r'\d+', f.stem)),
                     default=-1) + 1

    # Pre-load existing prompts into dedup so we never collide
    for bf in sorted(OUTPUT_DIR.glob("batch_*.json")):
        try:
            with open(bf, encoding="utf-8") as f:
                for item in json.load(f):
                    if "user_prompt" in item:
                        dedup.add(item["user_prompt"])
        except Exception:
            pass

    # Flatten todo into individual (component, operation) work items
    # Each pair may need > 1 example, so repeat accordingly
    work_items: List[Tuple[str, str]] = []
    for comp, op, needed in todo:
        work_items.extend([(comp, op)] * needed)

    random.shuffle(work_items)   # mix so batches aren't all the same component

    spec_id_counter = 0
    pending_specs: List[Dict] = []

    def flush_batch(specs):
        nonlocal batch_num
        print(f"\n[Batch {batch_num}] {len(specs)} items")
        results = generate_batch(client, specs, dedup)

        if results:
            print(f"  ✓ {len(results)} prayers generated")
            all_new.extend(results)
            for r in results:
                pair_counts[(r["component"], r["operation"])] += 1
            # Save batch file
            bf = OUTPUT_DIR / f"batch_{batch_num:04d}.json"
            with open(bf, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
        else:
            print(f"  ✗ Batch failed")

        batch_num += 1
        time.sleep(RATE_LIMIT_DELAY)

    for comp, op in work_items:
        prompt = make_prompt(comp, op)
        # Deduplicate prompts (not pairs — same pair can have 2 different prompts)
        attempts = 0
        while dedup.is_duplicate(prompt) and attempts < 20:
            prompt = make_prompt(comp, op)
            attempts += 1

        if dedup.is_duplicate(prompt):
            continue   # give up on this one

        dedup.add(prompt)
        pending_specs.append({
            "id":          f"s{spec_id_counter:05d}",
            "component":   comp,
            "operation":   op,
            "user_prompt": prompt,
        })
        spec_id_counter += 1

        if len(pending_specs) >= BATCH_SIZE:
            flush_batch(pending_specs)
            pending_specs = []

    if pending_specs:
        flush_batch(pending_specs)

    # ── Combine all results ───────────────────────────────────────────────────
    all_prayers = []
    for bf in sorted(OUTPUT_DIR.glob("batch_*.json")):
        try:
            with open(bf, encoding="utf-8") as f:
                all_prayers.extend(json.load(f))
        except Exception:
            pass

    combined_file = OUTPUT_DIR / "all_prayers.json"
    with open(combined_file, "w", encoding="utf-8") as f:
        json.dump(all_prayers, f, indent=2, ensure_ascii=False)

    training_file = OUTPUT_DIR / "training_data.txt"
    with open(training_file, "w", encoding="utf-8") as f:
        for item in all_prayers:
            f.write(f"<|user|>{item['user_prompt']}<|end|>\n")
            f.write(f"<|assistant|>{item['prayer']}<|end|>\n\n")

    # ── Update Excel pair log ─────────────────────────────────────────────────
    write_pair_log(EXCEL_FILE, pair_counts, components, operations, TARGET_PER_PAIR)

    # ── Final stats ───────────────────────────────────────────────────────────
    pairs_done_now = sum(1 for c in components for o in operations
                         if pair_counts.get((c, o), 0) >= TARGET_PER_PAIR)

    print("\n" + "=" * 70)
    print("✓ GENERATION COMPLETE")
    print("=" * 70)
    print(f"New prayers this run:  {len(all_new)}")
    print(f"Total in dataset:      {len(all_prayers)}")
    print(f"Pairs at target:       {pairs_done_now}/{pairs_total} "
          f"({100*pairs_done_now/pairs_total:.1f}%)")
    print(f"\nFiles written:")
    print(f"  {combined_file}")
    print(f"  {training_file}")
    print(f"  {EXCEL_FILE}  (Pair Log updated)")


if __name__ == "__main__":
    main()
